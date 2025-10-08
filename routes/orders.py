from flask import Blueprint, request, render_template, redirect
from openpyxl import load_workbook, Workbook
import os

orders_bp = Blueprint('orders', __name__, url_prefix='/orders')
FILE_PATH = "data.xlsx"
ORDERS_SHEET = "Orders"

# 1️⃣ Create Excel file safely if missing (shared with customers)
def create_excel_file():
    if os.path.exists(FILE_PATH):
        return  # Already created
    wb = Workbook()
    # Customers sheet
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.append(["Name", "Email", "Phone"])
    # Products sheet
    ws2 = wb.create_sheet("Products")
    ws2.append(["Name", "Description", "Price"])
    # Orders sheet
    ws3 = wb.create_sheet("Orders")
    ws3.append(["Customer", "Product", "Quantity"])
    wb.save(FILE_PATH)

# 2️⃣ Add order safely
def add_order_to_excel(customer, product, quantity):
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[ORDERS_SHEET]
    ws.append([customer, product, quantity])
    wb.save(FILE_PATH)

# 3️⃣ Load orders
def load_orders():
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[ORDERS_SHEET]
    data = []
    rows = list(ws.rows)
    headers = [cell.value for cell in rows[0]]
    for row in rows[1:]:
        data.append({headers[i]: row[i].value for i in range(len(headers))})
    return data

# 4️⃣ Flask routes
@orders_bp.route('/place', methods=['GET','POST'])
def place_order():
    if request.method == 'POST':
        customer = request.form.get('customer')
        product = request.form.get('product')
        quantity = request.form.get('quantity')
        if not customer or not product or not quantity:
            return "Fill all fields"
        add_order_to_excel(customer, product, quantity)
        return redirect('/orders/view')
    
     # Load customers and products for dropdown
    wb = load_workbook(FILE_PATH)
    customers_ws = wb["Customers"]
    products_ws = wb["Products"]

     # Convert to list
    customers = [row[0].value for row in list(customers_ws.rows)[1:]]
    products = [row[0].value for row in list(products_ws.rows)[1:]]
    return render_template('place_order.html', customers=customers, products=products)

@orders_bp.route('/view')
def view_orders():
    orders = load_orders()
    return render_template('view_orders.html', orders=orders)
