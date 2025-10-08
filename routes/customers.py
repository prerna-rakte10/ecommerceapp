from flask import Blueprint, request, render_template, redirect
from openpyxl import Workbook, load_workbook
import os

customers_bp = Blueprint('customers', __name__, url_prefix='/customers')
FILE_PATH = "data.xlsx"
CUSTOMERS_SHEET = "Customers"

# -----------------------------
# 1️⃣ Create Excel safely
# -----------------------------
def create_excel_file():
    if os.path.exists(FILE_PATH):
        return  # File already exists
    wb = Workbook()
    # Customers sheet
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.append(["Name", "Email", "Phone"])
    # Products sheet
    ws2 = wb.create_sheet("Products")
    ws2.append(["Name", "Description", "Price"])
    # Orders sheet
    ws3 = wb.create_sheet("Orders")
    ws3.append(["Customer", "Product", "Quantity"])
    wb.save(FILE_PATH)

# -----------------------------
# 2️⃣ Add customer safely
# -----------------------------
def add_customer_to_excel(name, email, phone):
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[CUSTOMERS_SHEET]
    ws.append([name, email, phone])
    wb.save(FILE_PATH)

# -----------------------------
# 3️⃣ Load customers to display
# -----------------------------
def load_customers():
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[CUSTOMERS_SHEET]
    data = []
    rows = list(ws.rows)
    headers = [cell.value for cell in rows[0]]
    for row in rows[1:]:
        data.append({headers[i]: row[i].value for i in range(len(headers))})
    return data

# -----------------------------
# 4️⃣ Flask routes
# -----------------------------
# Add customer
@customers_bp.route('/add', methods=['GET','POST'])
def add_customer():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        if not name or not email or not phone:
            return "Fill all fields"
        add_customer_to_excel(name, email, phone)
        return redirect('/customers/view')
    return render_template('add_customer.html')

# View customers
@customers_bp.route('/view')
def view_customers():
    customers = load_customers()
    return render_template('view_customers.html', customers=customers)

# -----------------------------
# 5️⃣ Edit customer
# -----------------------------
@customers_bp.route('/edit/<int:index>', methods=['GET','POST'])
def edit_customer(index):
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[CUSTOMERS_SHEET]
    rows = list(ws.rows)
    if request.method == 'POST':
        ws.cell(row=index+2, column=1).value = request.form['name']
        ws.cell(row=index+2, column=2).value = request.form['email']
        ws.cell(row=index+2, column=3).value = request.form['phone']
        wb.save(FILE_PATH)
        return redirect('/customers/view')
    # Pre-fill form with current customer data
    row_data = [cell.value for cell in rows[index+1]]
    return render_template('edit_customer.html', customer=row_data, index=index)

# -----------------------------
# 6️⃣ Delete customer
# -----------------------------
@customers_bp.route('/delete/<int:index>')
def delete_customer(index):
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[CUSTOMERS_SHEET]
    ws.delete_rows(index+2)  # +2 because row 1 is headers
    wb.save(FILE_PATH)
    return redirect('/customers/view')
