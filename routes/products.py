from flask import Blueprint, request, render_template, redirect
from openpyxl import load_workbook, Workbook
import os

products_bp = Blueprint('products', __name__, url_prefix='/products')
FILE_PATH = "data.xlsx"
PRODUCTS_SHEET = "Products"

# 1️⃣ Create Excel file safely if missing (shared with customers)
def create_excel_file():
    if os.path.exists(FILE_PATH):
        return  # Already created by customers
    wb = Workbook()
    # Customers sheet
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.append(["Name", "Email", "Phone"])
    # Products sheet
    ws2 = wb.create_sheet("Products")
    ws2.append(["Name", "Description", "Price"])
    # Orders sheet
    ws3 = wb.create_sheet("Orders")
    ws3.append(["Customer", "Product", "Quantity"])
    wb.save(FILE_PATH)

# 2️⃣ Add product safely
def add_product_to_excel(name, description, price):
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[PRODUCTS_SHEET]
    ws.append([name, description, price])
    wb.save(FILE_PATH)

# 3️⃣ Load products
def load_products():
    if not os.path.exists(FILE_PATH):
        create_excel_file()
    wb = load_workbook(FILE_PATH)
    ws = wb[PRODUCTS_SHEET]
    data = []
    rows = list(ws.rows)
    headers = [cell.value for cell in rows[0]]
    for row in rows[1:]:
        data.append({headers[i]: row[i].value for i in range(len(headers))})
    return data

# 4️⃣ Flask routes
@products_bp.route('/add', methods=['GET','POST'])
def add_product():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        price = request.form.get('price')
        if not name or not description or not price:
            return "Fill all fields"
        add_product_to_excel(name, description, price)
        return redirect('/products/view')
    return render_template('add_product.html')

@products_bp.route('/view')
def view_products():
    products = load_products()
    return render_template('view_products.html', products=products)
